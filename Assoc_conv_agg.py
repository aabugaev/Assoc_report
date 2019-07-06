import subprocess
import sys


def install(package):
	subprocess.call([sys.executable, "-m", "pip", "install", package])


try:
	import pandas as pd
	import numpy as np
except:
	install('pandas')
	install('numpy')


import pandas as pd
import numpy as np
import os
import re
FileList = os.listdir()


# In[186]:


csv_files = list(filter(lambda x: x.endswith("csv"), FileList ))

csv_df = [pd.read_csv(csv_file,comment= "#", encoding = "UTF-8" ) for csv_file in csv_files ]
Total = pd.concat(csv_df)


# In[187]:


Total["Ценность ассоциированных конверсий"] = Total["Ценность ассоциированных конверсий"].str.replace(',','.')
Total["Ценность ассоциированных конверсий"] = Total["Ценность ассоциированных конверсий"].str.replace('\s+|₽','')

cols = ["Ассоциированные конверсии","Ценность ассоциированных конверсий"]

Total[cols] = Total[cols].apply(pd.to_numeric, errors='coerce')
Total_Selected = Total.loc[:,["Источник или канал", "Кампания","Ассоциированные конверсии", "Ценность ассоциированных конверсий"]]

by_channel =Total_Selected.groupby("Источник или канал").sum()
by_channel_and_campaign = Total_Selected.groupby(["Источник или канал", "Кампания"]).sum()


# In[188]:


by_channel.to_excel("Assoc_by_channel.xlsx", merge_cells=False)
by_channel_and_campaign.to_excel("Assoc_by_channel_and_campaign.xlsx", merge_cells=False)


# In[189]:


"""
#%load file.py

===openpyxl===
from openpyxl import load_workbook
from openpyxl import Workbook

wb = load_workbook()
wb_ws = wb.get_active_sheet()

wrwb = Workbook()
wrwb_ws = wrwb.get_active_sheet()

wb.save()

===numpy/pandas===
import pandas as pd
import numpy as np

excel_df = pd.read_excel()
csv_df = pd.read_csv()


df.to_excel()
df.to_csv()

writer = pd.ExcelWriter('',engine='xlsxwriter',options={})
df.to_excel(writer)
writer.save()


====requests/BeautifulSoup===
import requests
from bs4 import BeautifulSoup

page = requests.get("http://yandex.ru")
page.encoding = "windows-1251"
soup = BeautifulSoup(''.join(page.text), "html.parser\"),
soup.findAll("div")


===Files and directories===
import os
FileList = os.listdir()

"""

